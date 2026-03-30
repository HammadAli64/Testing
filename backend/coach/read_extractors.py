"""
Extract plain text from common document formats for RAG ingestion.
Used by backend/read/ folder batch indexing.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from docx import Document  # type: ignore

READ_SUPPORTED_SUFFIXES = frozenset(
    {".docx", ".pdf", ".xlsx", ".csv", ".json", ".txt", ".md"},
)


def extract_docx_text(path: str) -> str:
    """
    Preserve Word heading styles as markdown-style # headings so RAG chunks
    and the planner can see document structure (titles / sections).
    """
    doc = Document(path)
    parts: list[str] = []
    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if not t:
            continue
        style_name = (p.style.name if p.style and p.style.name else "") or ""
        st = style_name.strip().lower()
        level = 0
        if "heading" in st:
            digits = "".join(ch for ch in st if ch.isdigit())
            if digits:
                level = min(max(int(digits), 1), 3)
            elif "1" in st or st.endswith("heading 1"):
                level = 1
            elif "3" in st or "4" in st:
                level = 3
            else:
                level = 2
        elif st in ("title",):
            level = 1
        elif st in ("subtitle",):
            level = 2
        if level:
            prefix = "#" * level + " "
            parts.append(f"{prefix}{t}")
        else:
            parts.append(t)
    return "\n\n".join(parts).strip()


def extract_pdf_text(path: str) -> str:
    from pypdf import PdfReader  # type: ignore

    reader = PdfReader(path)
    parts: list[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t.strip():
            parts.append(t)
    return "\n\n".join(parts).strip()


def extract_xlsx_text(path: str) -> str:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet in wb.worksheets:
            parts.append(f"## Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    parts.append("\t".join(cells))
    finally:
        wb.close()
    return "\n\n".join(parts).strip()


def extract_csv_text(path: str) -> str:
    lines: list[str] = []
    with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            lines.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(lines).strip()


def extract_json_text(path: str) -> str:
    with open(path, encoding="utf-8", errors="replace") as f:
        data = json.load(f)
    return json.dumps(data, ensure_ascii=False, indent=2)


def extract_plain_text(path: str) -> str:
    return Path(path).read_text(encoding="utf-8", errors="replace").strip()


def extract_text_from_path(path: Path | str) -> str:
    p = Path(path)
    suf = p.suffix.lower()
    if suf == ".docx":
        return extract_docx_text(str(p))
    if suf == ".pdf":
        return extract_pdf_text(str(p))
    if suf == ".xlsx":
        return extract_xlsx_text(str(p))
    if suf == ".csv":
        return extract_csv_text(str(p))
    if suf == ".json":
        return extract_json_text(str(p))
    if suf in {".txt", ".md"}:
        return extract_plain_text(str(p))
    raise ValueError(f"Unsupported file type: {suf}")


def list_supported_files(read_dir: Path) -> list[Path]:
    if not read_dir.is_dir():
        return []
    out: list[Path] = []
    for p in sorted(read_dir.iterdir()):
        if p.is_file() and p.suffix.lower() in READ_SUPPORTED_SUFFIXES:
            out.append(p)
    return out
